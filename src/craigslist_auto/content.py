from __future__ import annotations

import hashlib
import json
import random
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from .config import EXCEL_PATH, LOGS_DIR, PHOTOS_DIR

PHOTO_REUSE_LOG = LOGS_DIR / "photo_usage.json"
CONTENT_HASH_LOG = LOGS_DIR / "content_hashes.json"
PHOTO_COOLDOWN_DAYS = 30


@dataclass
class Ad:
    title: str
    body: str
    county: str
    city: str
    service_offered: str
    postal_code: str
    license_number: str
    phone_number: str
    photos: list[Path]
    source_row: int

    def content_hash(self) -> str:
        h = hashlib.sha256()
        h.update(self.title.encode())
        h.update(self.body.encode())
        return h.hexdigest()


def _load_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def _save_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")


def _expand_spintax(text: str, rng: random.Random) -> str:
    """Expand {a|b|c} → one of a/b/c. Nested supported."""
    pattern = re.compile(r"\{([^{}]+)\}")
    while True:
        m = pattern.search(text)
        if not m:
            return text
        choice = rng.choice(m.group(1).split("|"))
        text = text[: m.start()] + choice + text[m.end() :]


def _substitute_tokens(text: str, tokens: dict[str, str]) -> str:
    for k, v in tokens.items():
        text = text.replace("{" + k + "}", str(v))
    return text


def _load_excel_rows() -> list[dict]:
    """
    Excel schema (sheet 'ads'):
      county
      city
      service_offered
      posting_title  — supports spintax {a|b} and tokens like {city}, {zip_code}
      zip_code
      description    — same spintax/token support
      license_number
      phone_number
      photos_count   — optional, how many photos to attach (1-12). Blank → 1
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Excel file not found at {EXCEL_PATH}. "
            f"Run `uv run cl init-data` to create a sample."
        )
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["ads"] if "ads" in wb.sheetnames else wb.active
    raw_headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    headers = [_normalize_header(h) for h in raw_headers]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        rec["_row"] = i
        rows.append(rec)
    return rows


# Map any reasonable column name → the canonical key the code uses.
_HEADER_ALIASES = {
    "posting_title": "posting_title",
    "title": "posting_title",
    "description": "description",
    "body": "description",
    "post_body": "description",
    "city": "city",
    "county": "county",
    "service_offered": "service_offered",
    "service": "service_offered",
    "zip_code": "zip_code",
    "zip": "zip_code",
    "zipcode": "zip_code",
    "postal_code": "zip_code",
    "postal": "zip_code",
    "license_number": "license_number",
    "license": "license_number",
    "licensed": "license_number",
    "license_no": "license_number",
    "phone_number": "phone_number",
    "phone": "phone_number",
    "number": "phone_number",
    "photos_count": "photos_count",
    "photo_count": "photos_count",
    "photos": "photos_count",
}


def _normalize_header(h) -> str:
    if h is None:
        return ""
    key = str(h).strip().lower().replace(" ", "_").replace("-", "_")
    return _HEADER_ALIASES.get(key, key)


def _select_photos(account_photo_dir: Path, count: int, rng: random.Random) -> list[Path]:
    usage = _load_json(PHOTO_REUSE_LOG, {})
    now = datetime.now(timezone.utc)
    candidates = sorted(
        [p for p in account_photo_dir.iterdir() if p.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}]
    )
    if not candidates:
        raise RuntimeError(
            f"No photos found in {account_photo_dir}. "
            f"Add unique roofing photos for this account."
        )
    # Filter out recently used (within cooldown)
    fresh = []
    for p in candidates:
        last_used = usage.get(str(p))
        if last_used is None:
            fresh.append(p)
            continue
        last = datetime.fromisoformat(last_used)
        if (now - last).days >= PHOTO_COOLDOWN_DAYS:
            fresh.append(p)
    if len(fresh) < count:
        # Not enough fresh; allow oldest-used ones
        ordered = sorted(candidates, key=lambda p: usage.get(str(p), "0000"))
        fresh = ordered[: max(count, len(fresh))]
    rng.shuffle(fresh)
    return fresh[:count]


def mark_photos_used(photos: list[Path]) -> None:
    usage = _load_json(PHOTO_REUSE_LOG, {})
    now = datetime.now(timezone.utc).isoformat()
    for p in photos:
        usage[str(p)] = now
    _save_json(PHOTO_REUSE_LOG, usage)


def mark_content_used(ad: Ad) -> None:
    hashes = _load_json(CONTENT_HASH_LOG, [])
    hashes.append({"hash": ad.content_hash(), "at": datetime.now(timezone.utc).isoformat(), "row": ad.source_row})
    _save_json(CONTENT_HASH_LOG, hashes)


def _recent_hashes(days: int = 60) -> set[str]:
    hashes = _load_json(CONTENT_HASH_LOG, [])
    cutoff = datetime.now(timezone.utc).timestamp() - days * 86400
    out = set()
    for h in hashes:
        try:
            ts = datetime.fromisoformat(h["at"]).timestamp()
        except Exception:
            continue
        if ts >= cutoff:
            out.add(h["hash"])
    return out


def generate_ad(account_photo_dir: Path, seed: int | None = None) -> Ad:
    """Pick a random row, expand spintax + tokens, pick photos. Avoid recent duplicates."""
    rng = random.Random(seed)
    rows = _load_excel_rows()
    if not rows:
        raise RuntimeError("No rows in Excel.")
    recent = _recent_hashes()

    for _ in range(30):
        row = rng.choice(rows)
        city = row.get("city") or ""
        zip_code = str(row.get("zip_code") or "")
        tokens = {
            "city": city,
            "county": row.get("county") or "",
            "zip_code": zip_code,
            "postal_code": zip_code,
            "service": row.get("service_offered") or "",
            "phone": str(row.get("phone_number") or ""),
            "license": str(row.get("license_number") or ""),
        }
        title = _expand_spintax(_substitute_tokens(row["posting_title"], tokens), rng)
        body = _expand_spintax(_substitute_tokens(row["description"], tokens), rng)
        ad = Ad(
            title=title.strip(),
            body=body.strip(),
            county=row.get("county") or "",
            city=city,
            service_offered=row.get("service_offered") or "",
            postal_code=zip_code,
            license_number=str(row.get("license_number") or ""),
            phone_number=str(row.get("phone_number") or ""),
            photos=[],
            source_row=row["_row"],
        )
        if ad.content_hash() not in recent:
            break
    else:
        # Couldn't find unique content; just use the last attempt
        pass

    photo_count = row.get("photos_count") or 1
    ad.photos = _select_photos(account_photo_dir, int(photo_count), rng)
    return ad
